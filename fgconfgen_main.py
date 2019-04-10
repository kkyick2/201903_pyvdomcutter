#!/usr/bin/env python
# by kkyick2
# import pkg
import os
# import 3rd parties pkg
import openpyxl
# import project pkg
from tools import logger
import conf

NEXTLINE = '\r\n'


def cidr_to_netmask(cidr):
  cidr = int(cidr)
  mask = (0xffffffff >> (32 - cidr)) << (32 - cidr)
  return (str( (0xff000000 & mask) >> 24)   + '.' +
          str( (0x00ff0000 & mask) >> 16)   + '.' +
          str( (0x0000ff00 & mask) >> 8)    + '.' +
          str( (0x000000ff & mask)))


def write_row(outF, row):
    outF.write(row)


def xls2_policy_list(infile, sheet, vdom):
    """
        @param:	infile: full path of the xlsx, accept two format
                    1) if have vdom column, filter the input para vdom after output
                    2) if no vdom column, assume all policy in same vdom
                sheet: xlsx sheet name
                vdom: for filter only this vdom to output list
        @return:
                policy_list:
                order_keys:
    """
    logger1 = logger.logger().get()
    # workbook object is created
    wb_obj = openpyxl.load_workbook(infile)
    sheet_obj = wb_obj.get_sheet_by_name(sheet)
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    print('start: p_xls2dictList: (vdom, sheet): {},{}'.format(vdom, sheet))
    logger1.info('start: p_xls2dictList: (vdom, sheet): {},{}'.format(vdom, sheet))
    logger1.info('***** vdom: all - raw sheet size with heading col,row: {},{}'.format(max_col, max_row))

    order_keys = []  # [vdom, id, name, ...]
    policy_elem = {}  # {vdom:root, id:1,name: policy1, ... }
    policy_list = []  # [{vdom:root, id:1,name: policy1, ... }, {vdom:root, id:2,name: policy2, ... }, ... }

    # loop for each row V in xlsx
    for y in range(1, max_row + 1):
        # loop for each column -> in xlsx
        for x in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=y, column=x)
            # 1st row is key
            if y is 1:
                order_keys.append(cell_obj.value)
            # 2nd row is content
            else:
                policy_key = order_keys[x - 1]
                policy_elem[policy_key] = cell_obj.value
                # logger1.debug('cell({}, {}): (key|val) is: {} | {}'.format(str(x),str(y),str(policy_key),str(cell_obj.value)))
        if len(policy_elem) > 0:
            # 1) if have vdom column, filter input vdom only for output
            if 'vdom' in order_keys:
                if policy_elem['vdom'] == vdom:
                    policy_list.append(policy_elem)
            # 2) if no vdom column, assume all policy in same vdom for output
            else:
                policy_list.append(policy_elem)
        # end of the policy, clear the dictionary and loop next one
        policy_elem = {}

    # after consolidate the sheet, result in policy_list
    logger1.info('***** {}, {} -- Total {} items'.format(vdom,sheet,len(policy_list)))
    print ('***** {}, {} -- Total {} items'.format(vdom,sheet,len(policy_list)))
    logger1.info('end')
    return policy_list, order_keys


def gen_conf_policy(dictList, keys, vdom):
    """
        @param:	policydictList: a list of policies in dictionary format
                    [ {'id' : '1', 'srcintf' : 'internal', ...}, {'id' : '2', 'srcintf' : 'external', ...}, ... ]
                keys: a list of unique seen keys
                    ['id', 'srcintf', 'dstintf', ...]
                vdom: vdom name for output "config vdom, edit <vdom>"
        @return: N/A
    """
    # output file name
    outfile = vdom + '_policy.txt'

    logger1 = logger.logger().get()
    print('start: gen_conf_policy: (vdom): {}, Total item to gen is {}'.format(vdom,len(dictList)))
    logger1.info('start: gen_conf_policy: (vdom): {}, Total item to gen is {}'.format(vdom,len(dictList)))
    print('***** output file {}'.format(outfile))
    if 'vdom' in keys:
        logger1.info('***** requirement xls vdom is: ' + str(dictList[0]['vdom']))
    else:
        logger1.info('***** requirement xls do not have vdom, skip vdom key')

    if dictList and keys:
        logger1.info('gen config to txt')
        with open(outfile, 'wb') as outF:
            # First section [config vdom, edit <vdom>, config firewall policy]
            write_row(outF, 'config vdom' + NEXTLINE)
            write_row(outF, 'edit ' + vdom + NEXTLINE)
            write_row(outF, 'config firewall policy' + NEXTLINE)
            write_row(outF,  NEXTLINE)
            # loop for each policy in dictionary format, {'vdom:'root,'id':'999,'ippool':None,'uuid':'948a...' .}
            for policydict in dictList:
                # loop for each key to follow the order
                for key in keys:
                    if key in policydict.keys() and policydict[key]:
                        # skip column in xlsx
                        if key == 'vdom':
                            continue
                        elif key == 'uuid':
                            continue
                        elif key == 'Config Change Date':
                            continue
                        # edit <id>
                        elif key == 'id':
                            write_row(outF, 'edit ' + str(policydict[key]) + NEXTLINE)
                        # <name> and <comments> line add ""
                        elif key == 'name' or key == 'comments':
                            # remove object's next line and space from xls
                            obj_list = str(policydict[key]).split()
                            string = ""
                            for obj in obj_list:
                                string = string + obj
                            write_row(outF, 'set ' + str(key) + ' "' + string + '"' + NEXTLINE)

                        elif key == 'srcaddr' or key == 'dstaddr' or key == 'service':
                            # remove object's next line and space from xls
                            obj_list = str(policydict[key]).split()
                            string = ""
                            for obj in obj_list:
                                if string is "":
                                    string = '"' + obj + '"'
                                else:
                                    string = string + ' "' + obj + '"'
                            write_row(outF, 'set ' + str(key) + ' ' + string + ''+ NEXTLINE)
                        else:
                            write_row(outF, 'set ' + str(key) + ' "' + str(policydict[key]) + '"' + NEXTLINE)
                # End of each policy [next]
                write_row(outF, "next" + NEXTLINE)
                write_row(outF, NEXTLINE)
            # End section [end]
            write_row(outF, "end" + NEXTLINE)
    logger1.info('end')
    return


def gen_conf_service(list, vdom):
    """
        @param:	list: a list of service object in List format
                vdom: vdom name for output "config vdom, edit <vdom>"
        @return: N/A
    """
    # output file name
    outfile = vdom + '_service.txt'

    logger1 = logger.logger().get()
    print('start: gen_conf_service: (vdom): {}, Total item to gen is {}'.format(vdom,len(list)))
    logger1.info('start: gen_conf_service: (vdom): {}, Total item to gen is {}'.format(vdom,len(list)))
    print('***** output file {}'.format(outfile))

    with open(outfile, 'wb') as outF:
        # First section [config vdom, edit <vdom>]
        write_row(outF, 'config vdom' + NEXTLINE)
        write_row(outF, 'edit ' + vdom + NEXTLINE)
        write_row(outF, 'config firewall service category' + NEXTLINE)
        write_row(outF, 'edit "HKEX_USE"' + NEXTLINE)
        write_row(outF, 'next' + NEXTLINE)
        write_row(outF, 'end' + NEXTLINE)
        write_row(outF, NEXTLINE)
        write_row(outF, 'config firewall service custom' + NEXTLINE)
        write_row(outF, NEXTLINE)
        # loop for each item in list
        for item in list:
            # edit "<service name>"
            # set category "HKEX_USE"
            write_row(outF, 'edit "' + str(item) + '"'+ NEXTLINE)
            write_row(outF, 'set category "HKEX_USE"' + NEXTLINE)

            # change to lower letter for checking, remarks: output is original letter
            item_lw = item.lower()

            # three types of service obj 1/tcp 2/udp 3/others
            if item_lw.startswith('tcp/'):
                #print 'tcp: ' + item
                write_row(outF, 'set tcp-portrange ' + item.split('/')[1] + NEXTLINE)

            elif item_lw.startswith('udp/'):
                #print 'udp: ' + item
                write_row(outF, 'set udp-portrange ' + item.split('/')[1] + NEXTLINE)

            else:
                print item + ' <-- !! Not standard object, need manuel edit'
                write_row(outF, 'set !!!!! need manuel edit !!!!!!! ' + NEXTLINE)

            # End of each policy [next]
            write_row(outF, "next" + NEXTLINE)
            write_row(outF, NEXTLINE)
        # End section [end]
        write_row(outF, 'end' + NEXTLINE)
    logger1.info('end')
    return


def gen_conf_address(list, vdom, addrDictList):
    """
        @param:	list: a list of address object in List format
                vdom: vdom name for output "config vdom, edit <vdom>"
                addrDictList: a List of Dict that contain associated-interface mapping
        @return: N/A
    """
    # output file name
    outfile = vdom + '_address.txt'

    logger1 = logger.logger().get()
    print('start: gen_conf_address: (vdom): {}, Total item to gen is {}'.format(vdom,len(list)))
    logger1.info('start: gen_conf_address: (vdom): {}, Total item to gen is {}'.format(vdom,len(list)))
    print('***** output file {}'.format(outfile))

    with open(outfile, 'wb') as outF:
        # First section [config vdom, edit <vdom>]
        write_row(outF, 'config vdom' + NEXTLINE)
        write_row(outF, 'edit ' + vdom + NEXTLINE)
        write_row(outF, 'config firewall address' + NEXTLINE)
        write_row(outF, NEXTLINE)
        # loop for each item in list
        for item in list:
            # edit "<address name>"
            write_row(outF, 'edit "' + str(item) + '"' + NEXTLINE)

            # set comment
            write_row(outF, 'set comment ' + '""' + NEXTLINE)

            # set associated-interface
            asso_int = 'TBC'
            for i in addrDictList:
                if item in i['addr']:
                    asso_int = i['asso_int']
            write_row(outF, 'set associated-interface ' + '"' + asso_int + '"' + NEXTLINE)

            # change to lower letter for checking, remarks: output is original letter
            item_lw = item.lower()

            # three types of service obj 1/Host 2/Net 3/Range 4/others
            if item_lw.startswith('host_'):
                #print 'Host: ' + item
                write_row(outF, 'set type ipmask' + NEXTLINE)
                subnet = item.split('_')[1]
                mask = '255.255.255.255'
                write_row(outF, 'set subnet ' + subnet + ' ' + mask + NEXTLINE)

            elif item_lw.startswith('net_'):
                #print 'Net: ' + item
                write_row(outF, 'set type ipmask' + NEXTLINE)
                # if format is Net_10.1.0.0/16, get the mask
                if '/' in item:
                    cidr = item.split('_')[1].split('/')[1]
                    subnet = item.split('_')[1].split('/')[0]
                    mask = cidr_to_netmask(cidr)
                    write_row(outF, 'set subnet ' + subnet + ' ' + mask + NEXTLINE)
                # if format is Net_10.1.0.0, assume mask is /24
                else:
                    subnet = item.split('_')[1]
                    mask = '255.255.255.0'
                    write_row(outF, 'set subnet ' + subnet + ' ' + mask + NEXTLINE)

            elif item_lw.startswith('range_'):
                #print 'Range: ' + item
                write_row(outF, 'set type iprange' + NEXTLINE)
                # eg Range_10.1.90.117-118 --> 10.1.90.117
                startip = item.split('_')[1].split('-')[0]
                # eg Range_10.1.90.117-118 --> 118
                endip_oct = item.split('_')[1].split('-')[1]
                # convent to 4 oct ['10', '1', '90', '117']
                ipoct = startip.split('.')
                endip = ipoct[0] + '.' + ipoct[1] + '.' + ipoct[2] + '.' + endip_oct
                write_row(outF, 'set start-ip ' + startip + NEXTLINE)
                write_row(outF, 'set end-ip ' + endip + NEXTLINE)

            else:
                print item + ' <-- !! Not standard object, need manuel edit'
                write_row(outF, 'set !!!!! need manuel edit !!!!!!! ' + NEXTLINE)
            # End of each policy [next]
            write_row(outF, "next" + NEXTLINE)
            write_row(outF, NEXTLINE)
        # End section [end]
        write_row(outF, 'end' + NEXTLINE)
    logger1.info('end')
    return


def xls2_obj_list(infile, sheet, vdom, key):
    """
        @param:	infile: full path of the xlsx, accept two format
                    1) if have vdom column, filter the input para vdom after output
                    2) if no vdom column, assume all policy in same vdom
                sheet: xlsx sheet name
                vdom: for filter only this vdom to output list
        @return:
                returnList
    """
    logger1 = logger.logger().get()
    print('start: p_getobjList: (vdom, sheet, key): {},{},{}'.format(vdom,sheet,key))
    logger1.info('start: p_getobjList: (vdom, sheet, key): {},{},{}'.format(vdom,sheet,key))

    # workbook object is created
    wb_obj = openpyxl.load_workbook(infile)
    sheet_obj = wb_obj.get_sheet_by_name(sheet)
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    order_keys = []  # [vdom, id, name, ...]
    policy_elem = {}  # {vdom:root, id:1,name: policy1, ... }
    policy_list = []  # [{vdom:root, id:1,name: policy1, ... }, {vdom:root, id:2,name: policy2, ... }, ... }

    # loop for each row V in xlsx
    for y in range(1, max_row + 1):
        # loop for each column -> in xlsx
        for x in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=y, column=x)
            # 1st row is key
            if y is 1:
                order_keys.append(cell_obj.value)
            # 2nd row is content
            else:
                policy_key = order_keys[x - 1]
                policy_elem[policy_key] = cell_obj.value
        if len(policy_elem) > 0:
            # 1) if have vdom column, filter input vdom only for output
            if 'vdom' in order_keys:
                if policy_elem['vdom'] == vdom:
                    policy_list.append(policy_elem)
            # 2) if no vdom column, assume all policy in same vdom for output
            else:
                policy_list.append(policy_elem)
        # end of the policy, clear the dictionary and loop next one
        policy_elem = {}
    # after consolidate the sheet, result in policy_list
    returnList = []
    for dict in policy_list:
        returnList.append(dict[key])
    logger1.info('***** {} {} -- Total: {} items'.format(vdom, sheet, len(returnList)))
    print('***** {} {} -- Total: {} items'.format(vdom, sheet, len(returnList)))
    logger1.info('end')
    return returnList


def flatten_list(l):
    """
        return a flatten List
    """
    return [y for x in l for y in x]


def remove_dup_list(l):
    """
        return a List with remove duplicate elements
    """
    returnList = []
    # Iterate over the original list and for each element
    # add it to uniqueList, if its not already there.
    for elem in l:
        if elem not in returnList:
            returnList.append(elem)
    return returnList


def raw2_uniq_list(key, req_polList):
    """
        @param:
                req_polList: a list of policies in dictionary format
                    [ {'id' : '1', 'srcintf' : 'internal', ...}, {'id' : '2', 'srcintf' : 'external', ...}, ... ]
                key: key of the target return list, eg, srcintf, dstintf
        @return
                uniqList:       Type1: List of the target key
                uniqDictList:  Type2: List of Dict of srcaddr/dstaddr
    """
    logger1 = logger.logger().get()
    print('start: raw2_uniq_list ' + key)
    logger1.info('start: raw2_uniq_list ' + key)

    # Type1: for other key, return in List
    rawList = [] # eg. ['Host_10.63.109.100', 'Host_10.63.109.101' ]

    # Type2: for srcaddr/srcaddr key, return in List of Dict
    rawDict = {} # eg. {'asso_int': u'E.TOF2', 'addr': 'Host_10.63.109.100'}
    rawDictList = [] # eg. [{'asso_int': 'E.TOF2', 'addr': 'Host_10.63.109.100'}, {'asso_int': 'E.TOF2', 'addr': 'Host_10.63.109.101'}...]

    # loop for each policy
    for dict in req_polList:
        # temp list to hold the key element, eg, addr objects
        temp = str(dict[key]).split()
        # print 'temp: ' + str(temp)
        for i in temp:
            # print 'i: ' + str(i)
            # Type1: append each element to the raw List
            rawList.append(i)
            # Type2: if the key is srcaddr/dstaddr, we have to map it's associated interface
            if 'srcaddr' in key:
                rawDict['addr'] = i
                rawDict['asso_int'] = dict['srcintf']
                rawDictList.append(rawDict)
                rawDict = {}
                #print 'ykkList: ' + str(ykk_List)
                #print 'rawList: ' + str(req_rawList)
            elif 'dstaddr' in key:
                rawDict['addr'] = i
                rawDict['asso_int'] = dict['dstintf']
                rawDictList.append(rawDict)
                rawDict = {}
                #print 'ykkList: ' + str(ykk_List)
                #print 'rawList: ' + str(req_rawList)


    uniqDictList = remove_dup_list(rawDictList)
    uniqList = remove_dup_list(rawList)
    logger1.debug('*****List: {}, b4 remove duplicate: {} in xls, after remove duplicate: {}'.format(key,len(rawList),len(uniqList)))
    logger1.debug('*****Dict: {}, b4 remove duplicate: {} in xls, after remove duplicate: {}'.format(key,len(rawDictList),len(uniqDictList)))
    # print('*****List: {}, b4 remove duplicate: {} in xls, after remove duplicate: {}'.format(key,len(rawList),len(uniqList)))
    # print('*****Dict: {}, b4 remove duplicate: {} in xls, after remove duplicate: {}'.format(key,len(rawDictList),len(uniqDictList)))
    logger1.info('end')
    return uniqList, uniqDictList


def analyze_bas_mis_obj(key, bas_l, req_l):
    """
        @param:
                key: key of the operation string, eg,service
                bas_l: baseline obj in List
                req_l: requirement obj in List
        @return
                missingObjList: find the requirement missing obj in baseline, return in list
    """
    logger1 = logger.logger().get()
    print('--- [' + key + '] ----------------------------------------------------------')
    print('start: analyze_bas_mis_obj: (key,bas_len,req_len) is ({},{},{})'.format(key,len(bas_l),len(req_l)))
    logger1.info('start: analyze_bas_mis_obj: (key,bas_len,req_len) is ({},{},{})'.format(key,len(bas_l),len(req_l)))

    missingObjList = []
    for i in req_l:
        if i in bas_l:
            continue
        else:
            missingObjList.append(i)
            print 'Item not in baseline: {}'.format(i)

    # analyze result
    # logger1.debug('baseline obj: {}'.format(bas_List))
    # logger1.debug('reqiremt obj: {}'.format(req_uniqList))
    logger1.info('*** No of obj in (key,baseline,req,missing) is ({},{},{},{})'.format(key,len(bas_l),len(req_l),len(missingObjList)))
    logger1.info('end')
    # print('*** baseline obj: {}'.format(bas_List))
    # print('*** reqiremt obj: {}'.format(req_uniqList))
    print('*** No of obj in (key,baseline,req,missing) is ({},{},{},{})'.format(key,len(bas_l),len(req_l),len(missingObjList)))
    print('-------------------------------------------------------------------------------')
    print('')
    return missingObjList


def analyze_bas_mis_route(bas_l, req_l):
    return


def start(base, req, vdom):

    # define logger
    logger1 = logger.logger().get()
    logger1.info('start script: '+__name__)

    # arg1/baseline, xls
    bas_conf = base
    bas_conf_sheet = '07policy'

    # arg2-4/gen conf, xls requirement, vdom name
    req_conf = req
    req_conf_sheet = 'Policy'
    vdom = vdom

    # define full path
    bas_conf_path = os.path.join(conf.FGCONFGEN_BAS_PATH, bas_conf)
    req_conf_path = os.path.join(conf.FGCONFGEN_REQ_PATH, req_conf)
    print('==================================================================================')
    print('====================Begin: Script argument input =================================')
    print('==================================================================================')
    print('=== baseline conf: {}'.format(bas_conf))
    print('=== requiremt xls: {}'.format(req_conf))
    print('=== requiremt tab: {}'.format(req_conf_sheet))
    print('=== doing vdom: {}'.format(vdom))
    print('')

    # A---prepare checking List
    # baseline, xls to List elements
    print('==================================================================================')
    print('================== Task A1: Baseline to list =====================================')
    bas_zonList = xls2_obj_list(bas_conf_path, '02zone', vdom, 'name')
    bas_rouList = xls2_obj_list(bas_conf_path, '03route', vdom, 'dst')
    bas_addList = xls2_obj_list(bas_conf_path, '04addr', vdom, 'name')
    bas_serList = xls2_obj_list(bas_conf_path, '06service', vdom, 'name')
    print('')

    # requirement, xls to List elements
    print('=================== Task A2: Requirement to list =================================')
    req_polList, keys = xls2_policy_list(req_conf_path, req_conf_sheet, vdom)
    print('')

    # B---analyze baseline missing obj
    print('==================================================================================')
    print('=================== Task B: Find missing object in baseline ======================')
    # ---------- task for service
    key = 'service'
    l, d = raw2_uniq_list(key, req_polList)
    missing_service = analyze_bas_mis_obj(key, bas_serList, l)

    # ---------- task for address
    key = 'srcaddr'
    l_srcaddr, d_srcaddr = raw2_uniq_list(key, req_polList)
    key = 'dstaddr'
    l_dstaddr, d_dstaddr = raw2_uniq_list(key, req_polList)

    key = 'srcaddr_and_dstaddr'
    addrList = l_srcaddr + l_dstaddr
    addrDictList = d_srcaddr + d_dstaddr
    # srcaddr_and_dstaddr may have duplicate items, remove them
    addrList = remove_dup_list(addrList)
    addrDictList = remove_dup_list(addrDictList)
    missing_address = analyze_bas_mis_obj(key, bas_addList, addrList)

    # ---------- task for zone
    key = 'srcintf'
    l_srcintf, d = raw2_uniq_list(key, req_polList)
    key = 'dstintf'
    l_dstintf, d = raw2_uniq_list(key, req_polList)
    key = 'zone'
    intfList = l_srcintf + l_dstintf
    # srcintf and dstintf may have duplicate items
    intfList = remove_dup_list(intfList)
    missing_zone = analyze_bas_mis_obj(key, bas_zonList, intfList)
    print('')

    # analyze baseline missing route
    print('=================== Task B: Find missing static route in baseline ================')
    key = 'dst'
    print missing_address

    print('')
    # C --- gen conf, policy
    print('===================================================================================')
    print('=================== Task C1: gen policy conf ======================================')
    gen_conf_policy(req_polList, keys, vdom)
    print('')

    # 4/gen conf, missing service
    print('=================== Task C2: gen missing service conf ============================')
    gen_conf_service(missing_service, vdom)
    print('')

    # 5/gen conf, missing address
    print('=================== Task C3: gen missing address conf ============================')
    gen_conf_address(missing_address, vdom, addrDictList)
    print('')

    logger1.info('end: '+__name__)


if __name__ == "__main__":

    bas_conf = 'CTFW03_20190324_0727.conf.xlsx'
    req_conf = 'CT03CASH2_20190328a.xlsx'
    vdom = 'CT03CASH2'

    start(bas_conf, req_conf, vdom)